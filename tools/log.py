# -*- coding: utf-8 -*-
#python3
import os
import time
import re
import sys
import random,time
import threading
import multiprocessing
import openpyxl
import configparser
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook


file='e:/项目/地理编码/LOG/logs/system.log.2018-12-30.1'
dir="e:/项目/地理编码/LOG/logs/"
filelist=[]
filename = os.path.basename(file).split('.')[0]
result1 = filename + '.txt'
result2 = filename + '.csv'
fileresult1 = dir + result1
fileresult2 = dir + result2

def log(file):
    sum=0
    total=0
    now=time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
    result='result'+now+'.txt'
    p=open(result,'w+',encoding='utf-8')
    with open(file,'r',encoding='utf-8') as f:
        for line in f.readlines():
            if "address" in line and "city" in line:
                r1 = re.findall('address":"\S+",',line)
                r2 = re.findall('city":"\S+",', line)
                if r1 :
                    q1=r1[0].split(':')[1]
                    q=q1.split(',')[0].strip('"')
                    if r2:
                        city1=r2[0].split(':')[1]
                        city=city1.split(',')[0].strip('"')
                    else:
                        city=""
                    print('\n')
                    p.write(q)
                    p.write(",")
                    p.write(city)
                    p.write('\n')

            else:
                continue
        p.close()

def geodir(dir):
    for root, dirs, files in os.walk(dir):
        for file in files:
            filelist.append(os.path.join(root, file))
    sum=0
    total=0
    print(filelist)
    now=time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
    result='result'+now+'.txt'
    p=open(result,'w+',encoding='utf-8')
    for file in filelist:
        with open(file,'r',encoding='utf-8') as f:
            for line in f.readlines():
                if "address" in line and "city" in line:
                    r1 = re.findall('address":"\S+",', line)
                    r2 = re.findall('city":"\S+",', line)
                    if r1:
                        q1 = r1[0].split(':')[1]
                        q = q1.split(',')[0].strip('"')
                        if r2:
                            city1 = r2[0].split(':')[1]
                            city = city1.split(',')[0].strip('"')
                        else:
                            city = ""
                        print('\n')
                        p.write(q)
                        p.write(",")
                        p.write(city)
                        p.write('\n')


def ex(file,sheet_name):
    #now=time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))

    p=open(fileresult1,'w+',encoding='utf-8')
    workbook_ = openpyxl.load_workbook(file) #导入工作表
    #sheetnames =workbook_.get_sheet_names() #获得表单名字
    # sheet = workbook_.get_sheet_by_name(sheetnames[sheet_name]) #从工作表中提取某一表单
    sheet = workbook_.get_sheet_by_name(sheet_name)
    for rowNum in range(1,sheet.max_row):
        for colNum in range(1, sheet.max_column + 1):
            if colNum <sheet.max_column:
                xls=sheet.cell(row=rowNum, column=colNum).value
                print(xls)
                p.write(str(xls))
                p.write(',')
            else:
                p.write(str(xls))
        p.write('\n')
    p.close()

def replace(file):
    p = open(fileresult2, 'w+', encoding='utf-8')
    with open(fileresult1, 'r', encoding='utf-8') as g:
        for line in g.readlines():
            if "None" in line:
                p.write(line.replace('None', ''))
    p.close()

        #print(url_)#获得数据
    # print(tile)
    # print(url)


# ex(file,"白名单")
# replace(fileresult2)
geodir(dir)
