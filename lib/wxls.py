# import xlrd,xlwt
# from xlutils import copy
import random,time
import threading
import multiprocessing
import openpyxl
import configparser
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import time,os
from lib.test_abstract import TestAbstract
conf_url="e:/project/Interface_api-master/conf/base.conf"
conf_ak="e:/project/Interface_api-master/conf/ak.conf"
conf_db="e:/project/Interface_api-master/conf/db.conf"
conf_sql="e:/project/Interface_api-master/conf/db.conf"


# thread=[]
# # wb=openpyxl.load_workbook(filename='2017-04-08-12_03_30.xlsx',read_only=False)
# # ts=wb.create_sheet("test11")
# now = time.strftime('%Y-%m-%d-%H_%M_%S', time.localtime(time.time()))
#
# class xls(threading.Thread):
#     def __init__(self):
#         threading.Thread.__init__(self)
#         self.wt = openpyxl.Workbook()
#         self.ts = self.wt.create_sheet("test")
#
#
#     def xls(self):
#
#         for i in range(1,6):
#             # for j in range(1,3):
#             self.ts.cell(row =i, column = 1,value = random.randint(100, 200))
#             self.ts.cell(row=i, column=2, value=random.randint(100, 200))
#
#
#     def run(self):
#         self.xls()
#
#     def save(self):
#         self.wt.save(now + ".xlsx")
#
#
#
# for i in range(1):
#     t1=xls()
#     thread.append(t1)
#
# start=time.time()
# for i in thread:
#     print('%s start' %i)
#
#     i.start()
#
# for i in thread:
#     print('%s join' %i)
#     i.join()
# end=time.time()
#
# print("closing...")
# print("time=%s" %(end-start))
# t1.save()
# print("end")




# wt = openpyxl.Workbook()
# ts = wt.create_sheet("test22")
# nrows=300
# pro=1
#
# def xls(ts):
#     for i in range(1,nrows):
#         for j in range(1,10):
#             ts.cell(row=i, column=j, value=random.randint(100, 200))
#     wt.save(now + ".xlsx")

#从excel读取tile和每行数据，然后将title和每行数据组成字典形式，存入list，形成传参

tile=[]
data = []
exfile="e:/GPS测试(zhcs).xlsx"
def get_data(filename,sheet_name,line):
    url_=[]
    url=[]#申明list
    tile=[]
    
    workbook_ = openpyxl.load_workbook(filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    # sheet = workbook_.get_sheet_by_name(sheetnames[sheet_name]) #从工作表中提取某一表单
    sheet = workbook_.get_sheet_by_name(sheet_name)
    print ("Work Sheet Rows:", sheet.max_row)
    print(sheet.max_column)

    for rowNum in range(line,sheet.max_row+1):

        if rowNum==line:
            for colNum in range(1, sheet.max_column + 1):
                tile.append(sheet.cell(row=rowNum,column=colNum).value)

        else:
            url_ = []
            for colNum in range(1, sheet.max_column + 1):
                url_.append(sheet.cell(row=rowNum,column=colNum).value)
            data.append(url_)
        #print(url_)#获得数据
    # print(tile)
    # print(url)
    # for i in range(len(url)):
    #     data.append(dict(zip(tile,url[i-1])))
    
    return data

def rgeo(url,data):

    for i in data:
        x=i[1]
        y=i[2]

        kk = {'x': x, \
              'y': y, \
              'opt': '', \
              'ak': '42204f2dcbf94db4aabf977e4535708c'}

        testAbstract = TestAbstract()
        res = testAbstract.requestsget(url,kk)
        if 'result' in res.keys():
            if 'name' in res["result"].keys():
                name = res["result"]["name"]

        else:
            name="查询不到"
            
        print(name)
        i.append(name)

    return name

def wexcel (name):
    wb = Workbook()
    
    #sheet = wb.create('广深大80店明细_'+name, index=0)
    sheet = wb.active
    print(tile)
    sheet.append(tile)
    for i in data:
        sheet.append(i)  # 这一步就可以做到将12345插入到一行中。
    wb.save('GPS测试(zhcs)__'+'.xlsx')  # 保存文件，注意以xlsx为文件扩展名

#----------------------------------------------------------------------------------------------------------
# url="http://gis-int.intsit.sfdc.com.cn:1080/rgeo/api"
# line=1
# city=["Sheet1"]
# for i in city:
#     tile=[]
#     data = []
#     v=get_data(exfile,i,line)
#     k=rgeo(url,data)
#     wexcel(i)

#----------------------------------------------------------------------------------------------------------

# if __name__ == '__main__':

    # for i in range(pro):
    #     t1=multiprocessing.Process(target =xls, args=())
    #     thread.append(t1)
    #
    # s=time.time()
    # for i in thread:
    #     print('%s start' %i)
    #     i.start()
    #
    # for i in thread:
    #     print('%s join' %i)
    #     i.join()
    # end=time.time()
    #
    # print("closing...")
    # print("time=%s" %(end-s))
    # pool = multiprocessing.Pool(processes=3)
    # for i in range(4):
    #     msg = "hello %d" % (i)
    #     pool.apply_async(xls,(msg, ))  # 维持执行的进程总数为processes，当一个进程执行完毕后会添加新的进程进去
    #
    # pool.close()
    # pool.join()

    # wt.save(now + ".xlsx")
    # print("end")

def report(name,url, p, r):
    now = time.strftime('%Y-%m-%d-%H_%M_%S', time.localtime(time.time()))
    wb = Workbook()
    ws1 = wb.create_sheet(title=name)
    # 给第一行、第一列赋值，此处不同版本，起始行号可能不同
    ws1.cell(row=1, column=1).value = url
    print(len(p))
    for rowNum in range(1,len(p)+1):
        ws1.cell(row=rowNum, column=2).value = str(p[rowNum - 1])
        ws1.cell(row=rowNum, column=3).value = str(r[rowNum - 1])
    wb.save(name+ now + ".xlsx")

def reporttxt(name,url, p, r,type="get",n='test'):
    now = time.strftime('%Y-%m-%d-%H_%M_%S', time.localtime(time.time()))


    file=os.path.dirname(os.path.dirname(__file__))+ '/report'+'/'+name+"_"+now+".txt"
    #file=dir+ '/report'+'/'+name+"_"+now+".txt"

    with open(file,'w+',encoding='utf_8')as f:
        f.write(url)
        f.write('\n\n')
        for i in range(len(p)):
            temp = []
            f.write(str(i+1))
            f.write('\t')
            # f.write(n[i])
            f.write('\n')

            if type=="get":
                for key in p[i]:
                    # print (key,type(key))
                    # print (p[i],type(p[i]))
                    # print (p[i][key],type(p[i][key]))
                    temp.append(key+"="+p[i][key])
                parameters="&".join(temp)
                case=url+"?"+parameters
                f.write(case)
                f.write('\n\n')
                # f.write(str(p[i]))
                # f.write('\n')
                f.write(str(r[i]))
                f.write('\n\n\n')
            if type == "post":
                # print(p[i])
                # for key in p[i]:
                #     # print (key,type(key))
                #     # print (p[i],type(p[i]))
                #     # print (p[i][key],type(p[i][key]))
                #     temp.append(key + "=" + p[i][key])
                # parameters = "&".join(temp)
                # case = url + "?" + parameters
                f.write(url)
                f.write('\n')
                f.write("post:")
                f.write('\n\n')
                f.write(str(p[i]))
                f.write('\n\n')
                f.write("result:")
                f.write('\n\n')
                f.write(str(r[i]))
                f.write('\n\n\n')

def geturl(config):
    config=config
    #print (config)
    conf = configparser.ConfigParser()
    conf.read(conf_url,encoding="utf-8-sig")
    # lists_header = conf.sections()
    # print(lists_header)
    boolean = conf.has_section(config)
    # print(config,"is",boolean)
    if boolean:
        url=conf[config]['url']
        return url

def getak(config):
    config = config
        # print (config)
    conf = configparser.ConfigParser()
    conf.read(conf_ak, encoding="utf-8-sig")
        # lists_header = conf.sections()
        # print(lists_header)
    boolean = conf.has_section(config)
        # print(config,"is",boolean)
    if boolean:
        ak = conf[config]['ak']
        return ak
#print (geturl("geo"))

def getdb(config):
    db=[]
    #print (config)
    conf = configparser.ConfigParser()
    conf.read(conf_db,encoding="utf-8-sig")
    boolean = conf.has_section(config)
    if boolean:
        host=conf[config]['host']
        username = conf[config]['username']
        password = conf[config]['password']
        port = conf[config]['port']
        database = conf[config]['database']
        db.append(host)
        db.append(username)
        db.append(password)
        db.append(port)
        db.append(database)
    return db

def getsql(config):
        config = config
        # print (config)
        conf = configparser.ConfigParser()
        conf.read(conf_ak, encoding="utf-8-sig")
        # lists_header = conf.sections()
        # print(lists_header)
        boolean = conf.has_section(config)
        # print(config,"is",boolean)
        if boolean:
            ak = conf[config]['ak']
            return ak

    # lists_header = conf.sections()
    # print(lists_header)
    #return jira
