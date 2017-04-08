# import xlrd,xlwt
# from xlutils import copy
import random,time
import threading
import openpyxl

# wt=xlwt.Workbook()
# ts=wt.add_sheet('test',cell_overwrite_ok=True)
#
# for i in range(10):
#     ts.write(i,0,random.randint(100,200))
#
# wt.save("3.xls")
#
# rt=xlrd.open_workbook("3.xls")
# cp=copy.copy(rt)
# ct=cp.get_sheet(0)
# for i in range(10):
#     ct.write(i,1,random.randint(100,200))
#
# cp.save("3.xls")

# thread=[]
# wt = xlwt.Workbook()
# ts = wt.add_sheet('5', cell_overwrite_ok=True)
#
# now = time.strftime('%Y-%m-%d-%H_%M_%S', time.localtime(time.time()))
#
# class we(threading.Thread):
#     def __init__(self,id,ts):
#         threading.Thread.__init__(self)
#         self.id = id
#         self.ts=ts
#
#     def wre(self):
#         for i in range(1000):
#             for j in range(300):
#                 ts.write(i, j, random.randint(100, 200))
#
#
#     def run(self):
#         self.wre()
#
#
#
# for i in range(5):
#     t1=we(i,ts)
#     thread.append(t1)
#
# for i in thread:
#     print('%s start' %i)
#     i.start()
#
# for i in thread:
#     print('%s join' %i)
#     i.join()
#
# print("end")
# wt.save(now+".xls")


thread=[]
# wb=openpyxl.load_workbook(filename='2017-04-08-12_03_30.xlsx',read_only=False)
# ts=wb.create_sheet("test11")

wt = openpyxl.Workbook()
ts=wt.create_sheet("test")

now = time.strftime('%Y-%m-%d-%H_%M_%S', time.localtime(time.time()))

class we(threading.Thread):
    def __init__(self,id,ts):
        threading.Thread.__init__(self)
        self.id = id
        self.ts=ts

    def wre(self):
        for i in range(1,100):
            for j in range(1,30):
                ts.cell(row =i, column = j,value = random.randint(100, 200))


    def run(self):
        self.wre()



for i in range(5):
    t1=we(i,ts)
    thread.append(t1)

for i in thread:
    print('%s start' %i)
    i.start()

for i in thread:
    print('%s join' %i)
    i.join()

print("end")
wb.save(now+".xlsx")